from tkinter import *
from tkinter import messagebox
import openpyxl
import pathlib
import customtkinter as ctk  # Use import direto
from openpyxl import Workbook

# Setando a aparência padrão do sistema
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearence()
        self.todo_sistema()
        btn_submit = ctk.CTkButton(self, text="Salvar dados".upper(), command=self.submit, fg_color="#151", hover_color="#131")
        btn_submit.place(x=300, y=420)

        btn_clear = ctk.CTkButton(self, text="Limpar".upper(), command=self.clear, fg_color="#555", hover_color="#333")
        btn_clear.place(x=500, y=420)
        
        #Texts variaveis
        self.name_value = StringVar()
        self.contact_value = StringVar()
        self.age_value = StringVar()
        self.address_value = StringVar()
        # Entradas

        self.name_entry = ctk.CTkEntry(self, width=350,textvariable=self.name_value, font=(
            "Century Gothic bold", 16), fg_color="transparent")

        self.contact_entry = ctk.CTkEntry(self, width=200,textvariable=self.contact_value, font=(
            "Century Gothic bold", 16), fg_color="transparent")

        self.age_entry = ctk.CTkEntry(self, width=150,textvariable=self.age_value, font=(
            "Century Gothic bold", 16), fg_color="transparent")

        self.address_entry = ctk.CTkEntry(self, width=200,textvariable=self.address_value, font=(
            "Century Gothic bold", 16), fg_color="transparent")

        # ComboBox
        self.gender_combobox = ctk.CTkComboBox(
            self, values=["Masculino", "Feminino"], font=("Century Gothic bold", 14), width=150)
        self.gender_combobox.set("Masculino")

        # Entrada de OBS
        self.obs_entry = ctk.CTkTextbox(self, width=470, height=150, font=(
            "arial", 18), border_color="#aaa", border_width=2, fg_color="transparent")

        # Labels
        self.lb_name = ctk.CTkLabel(self, text="Nome completo: ", font=(
            "Century Gothic bold", 16), text_color=["#000", "#fff"])

        self.lb_contact = ctk.CTkLabel(self, text="Contato", font=(
            "Century Gothic bold", 16), text_color=["#000", "#fff"])

        self.lb_age = ctk.CTkLabel(self, text="Idade", font=(
            "Century Gothic bold", 16), text_color=["#000", "#fff"])

        self.lb_gender = ctk.CTkLabel(self, text="Gênero", font=(
            "Century Gothic bold", 16), text_color=["#000", "#fff"])

        self.lb_address = ctk.CTkLabel(self, text="Endereço", font=(
            "Century Gothic bold", 16), text_color=["#000", "#fff"])

        self.lb_obs = ctk.CTkLabel(self, text="Observações", font=(
            "Century Gothic bold", 16), text_color=["#000", "#fff"])
        

        self.change_apm("System")  

        # Defina uma aparência padrão
        self.lb_name.place(x=50, y=120)
        self.name_entry.place(x=50, y=150)
        self.lb_contact.place(x=450, y=120)

        self.contact_entry.place(x=450, y=150)
        self.lb_age.place(x=300, y=190)

        self.age_entry.place(x=300, y=220)
        self.lb_gender.place(x=500, y=190)

        self.gender_combobox.place(x=500, y=220)
        self.lb_address.place(x=50, y=190)
        self.address_entry.place(x=50, y=220)
        self.lb_obs.place(x=50, y=260)
        self.obs_entry.place(x=180, y=260)

   
    def layout_config(self):
        self.title("Sistema de Clientes")
        self.geometry("700x500")

    def appearence(self):
        self.lb_apm = ctk.CTkLabel(
            self, text="Tema", bg_color="transparent", text_color=['#000', "#fff"])
        self.lb_apm.place(x=50, y=430)
        self.opt_apm = ctk.CTkOptionMenu(
            self, values=["Ligth", "Dark", "System"], command=self.change_apm)
        self.opt_apm.place(x=50, y=460)

    def todo_sistema(self):
        frame = ctk.CTkFrame(self, width=700, height=50, corner_radius=0,
                             bg_color="teal", fg_color="teal")
        frame.place(x=0, y=10)
        title = ctk.CTkLabel(frame, text="Sistema de Clientes", font=(
            "Century Gothic bold", 24), text_color="#FFF").place(x=190, y=10)

        span = ctk.CTkLabel(self, text="Por favor, preencha todos os campos do formulário!", font=(
            "Century Gothic bold", 16), text_color=["#000", "#fff"])
        span.place(x=50, y=70)

        file_path = 'Clientes.xlsx'
        if not pathlib.Path(file_path).is_file():
        # Se o arquivo não existir, crie um novo arquivo Excel com um cabeçalho
            workbook = Workbook()
            sheet = workbook.active
            sheet['A1'] = "Nome completo"
            sheet['B1'] = "Contato"
            sheet['C1'] = "Idade"
            sheet['D1'] = "Genero"
            sheet['E1'] = "Endereco"
            sheet['F1'] = "Observacoes"
            workbook.save(file_path)


    def submit(self):
     
    # Pegando os dados dos entrys
     name = self.name_value.get()
     contact = self.contact_value.get()
     age = self.age_value.get()
     gender = self.gender_combobox.get()
     address = self.address_value.get()
     obs = self.obs_entry.get(1.0, END) 
      
     if name == "" or contact == "" or age == "" or address == "":
      messagebox.showerror("Sistema", "ERRO!\nPor favor, preencha todos os campos!")
     else:
      ficheiro = openpyxl.load_workbook('Clientes.xlsx')
      folha = ficheiro.active
      folha.cell(column=1, row=folha.max_row + 1, value=name)
      folha.cell(column=2, row=folha.max_row, value=contact)
      folha.cell(column=3, row=folha.max_row, value=age)
      folha.cell(column=4, row=folha.max_row, value=gender)
      folha.cell(column=5, row=folha.max_row, value=address)
      folha.cell(column=6, row=folha.max_row, value=obs)

     ficheiro.save("Clientes.xlsx")
     messagebox.showinfo("Sistema", "Dados salvos com sucesso!")


    def clear (self):
      self.name_value.set("")
      self.contact_value.set("")
      self.age_value.set("")
      self.address_value.set("")
      self.obs_entry.delete(1.0, END)


    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)


if __name__ == "__main__":
    app = App()
    app.mainloop()
